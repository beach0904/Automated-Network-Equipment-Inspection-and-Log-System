import sqlite3
import pandas as pd
import glob
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class NetworkMonitor:
    def __init__(self, db_name="NetworkOps.db"):
        self.db_name = db_name
        self.init_db()
    def init_db(self):
        """
        [v2.0.0 改版] 初始化純淨 3NF 資料庫
        完全對應 Kaggle 真實數據欄位，移除區域相依性
        """
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()

        # 1. Devices 表 (儲存設備名稱，實現 3NF)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS Devices (
                device_id INTEGER PRIMARY KEY AUTOINCREMENT,
                device_name TEXT UNIQUE NOT NULL
            )
        ''')

        # 2. PerformanceLogs 表 (完整對應圖 1 數據維度)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS PerformanceLogs (
                log_id INTEGER PRIMARY KEY AUTOINCREMENT,
                device_id INTEGER,
            
                -- 真實數據指標 (REAL 型態確保精準度)
                cpu_utilization REAL,
                memory_utilization REAL,
                disk_io REAL,
                network_latency REAL,
                process_count INTEGER,
                thread_count INTEGER,
                context_switch INTEGER,
                cache_miss REAL,
                temperature REAL,
                power_consumption REAL,
                uptime REAL,
                status INTEGER,
            
                -- 系統產出欄位
                diagnosis_result TEXT,
                inspect_time DATETIME,
            
                FOREIGN KEY (device_id) REFERENCES Devices(device_id)
            )
        ''')

        conn.commit()
        conn.close()
        print(f"✅ v2.0.0 純淨版 3NF 資料庫架構已建立。")
    
    def process_source_data(self, folder_path="source_data"):
        """
        [v2.0.0 改版] 讀取純淨 Kaggle 數據並寫入 3NF 資料庫
        """
        # 同時支援掃描 .xlsx 與 .csv 格式
        files = glob.glob(os.path.join(folder_path, "*.xlsx")) + glob.glob(os.path.join(folder_path, "*.csv"))
        
        if not files:
            print("⚠️ 找不到來源檔案，請檢查 source_data 資料夾。")
            return

        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()

        for file in files:
            print(f"🚀 正在處理 Kaggle 真實數據源：{os.path.basename(file)}...")

            # 根據檔案類型讀取
            if file.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)

            # --- 數據清洗 (Data Cleaning) ---
            # 真實數據常有缺失值，統一補 0 確保資料庫寫入正常
            df = df.fillna(0) 

            for _, row in df.iterrows():
                # --- 1. 設備處理 (Devices 表) ---
                # 真實數據若無名稱則自動生成編號，以維持 3NF 關聯
                d_name = row.get('SystemName', f"Device-{_ + 1}")
                
                cursor.execute("INSERT OR IGNORE INTO Devices (device_name) VALUES (?)", (d_name,))
                cursor.execute("SELECT device_id FROM Devices WHERE device_name = ?", (d_name,))
                device_id = cursor.fetchone()[0]

                # --- 2. 智慧診斷邏輯 (基於圖一數據) ---
                raw_status = row.get('status', 0)
                # 假設 status 0 為正常，其餘為異常 (這可根據數據集說明調整)
                is_down = True if raw_status != 0 else False
                
                cpu = row.get('cpu_utilization', 0)
                temp = row.get('temperature', 0)
                
                # 判定邏輯：狀態異常或溫度過高(>90)為嚴重，CPU過高(>80)為警告
                if is_down or temp > 90:
                    diagnosis = "CRITICAL"
                elif cpu > 80:
                    diagnosis = "WARNING"
                else:
                    diagnosis = "NORMAL"

                # --- 3. 寫入效能日誌 (PerformanceLogs 表 - 完整對接 12 指標) ---
                cursor.execute('''
                    INSERT INTO PerformanceLogs (
                        device_id, cpu_utilization, memory_utilization, disk_io, 
                        network_latency, process_count, thread_count, context_switch, 
                        cache_miss, temperature, power_consumption, uptime, 
                        status, diagnosis_result, inspect_time
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    device_id,
                    cpu,
                    row.get('memory_utilization', 0),
                    row.get('disk_io', 0),
                    row.get('network_latency', 0),
                    int(row.get('process_count', 0)),
                    int(row.get('thread_count', 0)),
                    int(row.get('context_switch', 0)),
                    row.get('cache_miss', 0),
                    temp,
                    row.get('power_consumption', 0),
                    row.get('uptime', 0),
                    int(raw_status),
                    diagnosis,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ))

        conn.commit()
        conn.close()
        print(f"✨ v2.0.0 純淨數據導入完成，資料庫：{self.db_name}")

    def generate_report(self):
        """
        [v2.0.0 改版] 透過 3NF 關聯產出純淨 Kaggle 指標報表
        對接圖一 12 個效能欄位並實作全列自動標色
        """
        conn = sqlite3.connect(self.db_name)
        # 根據新資料庫架構進行 SQL JOIN
        query = '''
            SELECT 
                d.device_name AS '設備名稱', 
                p.cpu_utilization AS 'CPU使用率', 
                p.memory_utilization AS '記憶體使用率',
                p.disk_io AS 'Disk_IO',
                p.network_latency AS '網路延遲',
                p.process_count AS '進程數',
                p.thread_count AS '執行緒數',
                p.cache_miss AS '快取缺失',
                p.temperature AS '設備溫度',
                p.power_consumption AS '功耗',
                p.uptime AS '運行時間',
                p.status AS '狀態碼',
                p.diagnosis_result AS '診斷結果', 
                p.inspect_time AS '巡檢時間'
            FROM PerformanceLogs p
            JOIN Devices d ON p.device_id = d.device_id
            WHERE p.diagnosis_result != 'NORMAL'
            ORDER BY p.inspect_time DESC
        '''
        error_df = pd.read_sql_query(query, conn)
        conn.close()
        
        if not error_df.empty:
            output_file = f"真實數據異常報告_{datetime.now().strftime('%Y%m%d')}.xlsx"
            
            # 匯出資料
            error_df.to_excel(output_file, index=False)
            
            # 使用 openpyxl 進行標色渲染
            wb = load_workbook(output_file)
            ws = wb.active
            
            # 定義標色樣式
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            orange_fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")

            # 遍歷資料列 (從第二列開始)
            # 在新欄位順序中，「診斷結果」目前在第 13 欄 (M 欄)
            diag_column_index = 13
            # 總共有 14 個欄位 (從設備名稱到巡檢時間)
            total_columns = 14

            for row in range(2, ws.max_row + 1):
                diagnosis = ws.cell(row=row, column=diag_column_index).value
                
                if diagnosis == "CRITICAL":
                    for col in range(1, total_columns + 1):
                        ws.cell(row=row, column=col).fill = red_fill
                elif diagnosis == "WARNING":
                    for col in range(1, total_columns + 1):
                        ws.cell(row=row, column=col).fill = orange_fill

            # 自動調整欄寬
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = max_length + 2

            wb.save(output_file)
            print(f"🚩 v2.0.0 真實數據報表已產出：{output_file}")
        else:
            print("🙌 所有數據指標均在正常範圍內。")

if __name__ == "__main__":
    monitor = NetworkMonitor()
    # 確保 source_data 資料夾存在
    if os.path.exists("source_data"):
        monitor.process_source_data()
        monitor.generate_report()
    else:
        print("⚠️ 錯誤：找不到 source_data 資料夾，請確保該路徑下有區域 Excel 檔案。")