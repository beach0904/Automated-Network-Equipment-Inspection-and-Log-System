import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os

def run_network_analysis(source_file):
    if not os.path.exists(source_file):
        print(f"錯誤：找不到來源檔案 {source_file}")
        return

    print(f"正在讀取模擬數據: {source_file}...")
    
    # 讀取來源資料
    df = pd.read_excel(source_file)
    
    # 執行自動化判斷邏輯
    # 規則：CPU > 80% 或 狀態為 Down 判定為異常
    def check_status(row):
        if row['介面狀態'] == 'Down':
            return 'CRITICAL (斷線)'
        elif row['CPU使用率(%)'] > 80:
            return 'WARNING (負載過高)'
        else:
            return 'NORMAL (正常)'

    df['診斷結果'] = df.apply(check_status, axis=1)
    
    # 產出最終分析報表
    output_file = f"網路維運分析報告_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    # 使用 ExcelWriter 配合 openpyxl 進行格式美化
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='分析結果')
    
    workbook  = writer.book
    worksheet = writer.sheets['分析結果']
    
    # 樣式定義
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    center_align = Alignment(horizontal='center')
    
    # 設定標題列樣式
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center_align

    # 根據診斷結果自動標色 (遍歷每一列)
    for row_idx in range(2, len(df) + 2):
        diagnosis = worksheet.cell(row=row_idx, column=7).value # 診斷結果在第 7 欄
        if 'CRITICAL' in diagnosis:
            for col_idx in range(1, 8):
                worksheet.cell(row=row_idx, column=col_idx).fill = red_fill
        elif 'WARNING' in diagnosis:
            for col_idx in range(1, 8):
                worksheet.cell(row=row_idx, column=col_idx).fill = yellow_fill

    # 自動調整欄寬
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        worksheet.column_dimensions[column_letter].width = max_length + 2

    writer.close()
    print(f"✅ 分析完成！已產出專業報告：{output_file}")

if __name__ == "__main__":
    # 這裡指定讀取你提供的模擬資料檔
    run_network_analysis('Network_Inventory_Mock_Data.xlsx')